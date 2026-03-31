"""Generate incremental daily data for Airflow ETL pipeline.

Produces new orders, clickstream events, and CS tickets for a given date,
then pushes them to the running source systems (PostgreSQL, FastAPI, Excel dir).

Price distribution is calibrated to match the real Olist dataset:
    - Median order value: ~R$107
    - Mean order value:   ~R$160
    - P75: ~R$180,  P95: ~R$475
    - Items per order:    ~1.0  (95% single-item)

Usage:
    python -m generate_fake_data.generate_daily_data --date 2026-03-27
    python -m generate_fake_data.generate_daily_data  # defaults to today
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import math
import os
import random
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("generate_daily")

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_SAMPLE_DIR = _PROJECT_ROOT / "sample_data"
_DATA_SOURCE_DIR = _PROJECT_ROOT / "data_source"

# ---------------------------------------------------------------------------
# Load reference data (user_ids, product_ids) from sample_data
# ---------------------------------------------------------------------------

def _load_unique_ids(filename: str, column: str) -> list[str]:
    path = _SAMPLE_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"Missing: {path}")
    ids: set[str] = set()
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            val = row.get(column, "").strip()
            if val:
                ids.add(val)
    return list(ids)


def _load_reference_data() -> tuple[list[str], list[str]]:
    customer_ids = _load_unique_ids("olist_customers_dataset.csv", "customer_unique_id")
    product_ids = _load_unique_ids("olist_products_dataset.csv", "product_id")
    logger.info("Reference data: %d customers, %d products", len(customer_ids), len(product_ids))
    return customer_ids, product_ids


# ---------------------------------------------------------------------------
# 1. Daily Orders → PostgreSQL (via SQL file or direct insert)
# ---------------------------------------------------------------------------

STATUS_OPTIONS = ["Pending", "Processing", "Completed"]
STATUS_WEIGHTS = [0.1, 0.15, 0.75]
PAYMENT_OPTIONS = ["Credit Card", "Bank Slip", "Voucher", "Debit Card"]
PAYMENT_WEIGHTS = [0.55, 0.20, 0.10, 0.15]


def _realistic_price(rng: random.Random) -> float:
    """Generate a price matching Olist's right-skewed distribution.

    Uses log-normal with mu=4.67, sigma=0.85 which gives:
        median ≈ R$107,  mean ≈ R$155,  P75 ≈ R$180,  P95 ≈ R$450
    Clamped to [10, 1500] to avoid extremes.
    """
    mu, sigma = 4.67, 0.85
    raw = math.exp(rng.gauss(mu, sigma))
    return round(max(10.0, min(raw, 1500.0)), 2)


def generate_daily_orders(
    rng: random.Random,
    target_date: datetime,
    customer_ids: list[str],
    product_ids: list[str],
    num_orders: int = 65,
) -> tuple[list[dict], list[dict]]:
    """Generate new orders and order_items for *target_date*.

    Default num_orders ≈ 65 simulates a mid-sized e-commerce (~2000 orders/month).
    Returns (orders, order_items).
    """
    orders: list[dict] = []
    items: list[dict] = []

    for _ in range(num_orders):
        order_id = uuid.uuid4().hex
        user_id = rng.choice(customer_ids)
        # Olist: 95%+ orders are single-item
        n_items = rng.choices([1, 2, 3], weights=[0.95, 0.04, 0.01], k=1)[0]
        total = 0.0

        # random time within the day
        hour = rng.randint(0, 23)
        minute = rng.randint(0, 59)
        second = rng.randint(0, 59)
        created_at = target_date.replace(hour=hour, minute=minute, second=second)

        for seq in range(1, n_items + 1):
            pid = rng.choice(product_ids)
            price = _realistic_price(rng)
            total += price
            items.append({
                "item_id": f"{order_id}_{seq}",
                "order_id": order_id,
                "product_id": pid,
                "quantity": 1,
                "unit_price": price,
            })

        status = rng.choices(STATUS_OPTIONS, weights=STATUS_WEIGHTS, k=1)[0]
        payment = rng.choices(PAYMENT_OPTIONS, weights=PAYMENT_WEIGHTS, k=1)[0]

        orders.append({
            "order_id": order_id,
            "user_id": user_id,
            "total_amount": round(total, 2),
            "order_status": status,
            "payment_method": payment,
            "created_at": created_at.strftime("%Y-%m-%d %H:%M:%S"),
        })

    logger.info("Generated %d orders, %d items for %s", len(orders), len(items), target_date.date())
    return orders, items


def write_daily_sql(orders: list[dict], items: list[dict], target_date: datetime) -> Path:
    """Write INSERT-only SQL file for daily orders."""
    output_dir = _PROJECT_ROOT / "infra" / "source-db" / "daily"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"daily_{target_date.strftime('%Y%m%d')}.sql"
    path = output_dir / filename

    def _esc(v: str) -> str:
        if not v:
            return "NULL"
        return "'" + v.replace("'", "''") + "'"

    with open(path, "w", encoding="utf-8") as f:
        f.write(f"-- Daily data for {target_date.date()}\n\n")

        # Orders
        if orders:
            cols = "order_id, user_id, total_amount, order_status, payment_method, created_at"
            f.write(f"INSERT INTO orders ({cols}) VALUES\n")
            lines = []
            for o in orders:
                vals = ", ".join([
                    _esc(o["order_id"]), _esc(o["user_id"]),
                    str(o["total_amount"]), _esc(o["order_status"]),
                    _esc(o["payment_method"]), _esc(o["created_at"]),
                ])
                lines.append(f"  ({vals})")
            f.write(",\n".join(lines) + ";\n\n")

        # Order items
        if items:
            cols = "item_id, order_id, product_id, quantity, unit_price"
            f.write(f"INSERT INTO order_items ({cols}) VALUES\n")
            lines = []
            for it in items:
                vals = ", ".join([
                    _esc(it["item_id"]), _esc(it["order_id"]),
                    _esc(it["product_id"]), str(it["quantity"]),
                    str(it["unit_price"]),
                ])
                lines.append(f"  ({vals})")
            f.write(",\n".join(lines) + ";\n")

    logger.info("Wrote daily SQL → %s", path)
    return path


# ---------------------------------------------------------------------------
# 2. Daily Clickstream Events → JSON file + POST to API
# ---------------------------------------------------------------------------

EVENT_TYPES = ["view_item", "add_to_cart", "cart_abandonment"]
DEVICE_OS_LIST = ["Android", "iOS", "Windows", "macOS", "Linux"]


def generate_daily_events(
    rng: random.Random,
    target_date: datetime,
    customer_ids: list[str],
    product_ids: list[str],
    num_sessions: int = 350,
) -> list[dict[str, Any]]:
    """Generate clickstream events for *target_date*."""
    events: list[dict] = []
    total_range = 24 * 3600  # seconds in a day

    for _ in range(num_sessions):
        offset = rng.random() * total_range
        session_start = target_date + timedelta(seconds=offset)
        session_id = str(uuid.uuid4())
        user_id = rng.choice(customer_ids) if rng.random() > 0.05 else None
        device = rng.choice(DEVICE_OS_LIST) if rng.random() > 0.03 else None
        current_time = session_start

        # 1-6 view_item
        for _ in range(rng.randint(1, 6)):
            events.append(_make_event(rng, product_ids, session_id, "view_item", current_time, user_id, device))
            current_time += timedelta(seconds=rng.randint(10, 120))

        # 40% → add_to_cart
        if rng.random() < 0.40:
            for _ in range(rng.randint(1, 3)):
                events.append(_make_event(rng, product_ids, session_id, "add_to_cart", current_time, user_id, device))
                current_time += timedelta(seconds=rng.randint(5, 60))

            # 35% → cart_abandonment
            if rng.random() < 0.35:
                events.append(_make_event(rng, product_ids, session_id, "cart_abandonment", current_time, user_id, device))

    logger.info("Generated %d events for %s", len(events), target_date.date())
    return events


def _make_event(
    rng: random.Random,
    product_ids: list[str],
    session_id: str,
    event_type: str,
    ts: datetime,
    user_id: str | None,
    device: str | None,
) -> dict[str, Any]:
    product_id = rng.choice(product_ids) if rng.random() > 0.05 else None
    time_spent = rng.randint(3, 300) if rng.random() > 0.02 else None
    return {
        "event_id": str(uuid.uuid4()),
        "timestamp": ts.strftime("%Y-%m-%dT%H:%M:%S.") + f"{rng.randint(0,999):03d}Z",
        "user_id": user_id,
        "session_id": session_id,
        "event_type": event_type,
        "product_id": product_id,
        "device_os": device,
        "time_spent_seconds": time_spent,
    }


def write_daily_events(events: list[dict], target_date: datetime) -> Path:
    """Write events to a date-partitioned JSON-lines file."""
    output_dir = _DATA_SOURCE_DIR / "api" / "daily"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"events_{target_date.strftime('%Y%m%d')}.json"
    path = output_dir / filename

    with open(path, "w", encoding="utf-8") as f:
        for event in events:
            f.write(json.dumps(event, ensure_ascii=False) + "\n")

    logger.info("Wrote %d events → %s", len(events), path)
    return path


def push_events_to_api(events: list[dict], api_url: str | None = None) -> None:
    """POST events to the FastAPI source server (best-effort)."""
    url = (api_url or os.getenv("SOURCE_API_URL", "http://localhost:8000")).rstrip("/")
    try:
        import requests
        resp = requests.post(f"{url}/events/batch", json=events, timeout=30)
        resp.raise_for_status()
        logger.info("Pushed %d events to API (%s)", len(events), url)
    except Exception as exc:
        logger.warning("Could not push events to API: %s", exc)


# ---------------------------------------------------------------------------
# 3. Daily CS Tickets → Excel
# ---------------------------------------------------------------------------

ISSUE_TYPES = ["Product Issue", "General Inquiry", "Positive Feedback"]
ISSUE_WEIGHTS = [0.25, 0.30, 0.45]


def generate_daily_tickets(
    rng: random.Random,
    target_date: datetime,
    customer_ids: list[str],
    num_tickets: int = 20,
) -> list[dict]:
    """Generate CS tickets for *target_date*."""
    from generate_fake_data.helpers import deterministic_email

    tickets: list[dict] = []
    for _ in range(num_tickets):
        uid = rng.choice(customer_ids)
        email = deterministic_email(uid)
        # ~5% dirty email
        if rng.random() < 0.05:
            email = email.replace("@", "")

        rating = rng.choices([1, 2, 3, 4, 5], weights=[0.08, 0.12, 0.20, 0.30, 0.30], k=1)[0]
        issue_type = "Product Issue" if rating <= 2 else ("General Inquiry" if rating == 3 else "Positive Feedback")
        status = rng.choices(["Open", "Resolved"], weights=[0.25, 0.75], k=1)[0]

        hour = rng.randint(8, 20)
        minute = rng.randint(0, 59)
        reported = target_date.replace(hour=hour, minute=minute).strftime("%Y-%m-%d %H:%M:%S")

        tickets.append({
            "Ticket_ID": str(uuid.uuid4()),
            "Order_ID": uuid.uuid4().hex,  # simplified — in real system would join to real order
            "Customer_Email": email,
            "Issue_Type": issue_type,
            "Status": status,
            "Customer_Rating": rating,
            "Reported_Date": reported,
        })

    logger.info("Generated %d tickets for %s", len(tickets), target_date.date())
    return tickets


def write_daily_tickets(tickets: list[dict], target_date: datetime) -> Path:
    """Write tickets to a date-partitioned XLSX file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    output_dir = _DATA_SOURCE_DIR / "excel"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"CS_Tickets_{target_date.strftime('%Y%m%d')}.xlsx"
    path = output_dir / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CS_Tickets"

    headers = ["Ticket_ID", "Order_ID", "Customer_Email", "Issue_Type",
               "Status", "Customer_Rating", "Reported_Date"]
    ws.append(headers)

    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = hfont
        cell.fill = hfill

    for t in tickets:
        ws.append([t[h] for h in headers])

    wb.save(path)
    logger.info("Wrote %d tickets → %s", len(tickets), path)
    return path


# ---------------------------------------------------------------------------
# Execute daily SQL against PostgreSQL
# ---------------------------------------------------------------------------

def execute_daily_sql(sql_path: Path) -> None:
    """Run the generated SQL file against the source PostgreSQL database."""
    import subprocess

    host = os.getenv("SOURCE_PG_HOST", "localhost")
    port = os.getenv("SOURCE_PG_PORT", "5432")
    db = os.getenv("SOURCE_PG_DB", "ecommerce")
    user = os.getenv("SOURCE_PG_USER", "source_user")
    password = os.getenv("SOURCE_PG_PASSWORD", "source_pass")

    env = {**os.environ, "PGPASSWORD": password}
    cmd = [
        "psql", "-h", host, "-p", port, "-U", user, "-d", db,
        "-f", str(sql_path),
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, env=env, timeout=60)
        if result.returncode == 0:
            logger.info("Executed SQL: %s", sql_path.name)
        else:
            logger.warning("psql stderr: %s", result.stderr[:500])
    except FileNotFoundError:
        logger.warning("psql not found — skipping DB insertion. SQL at: %s", sql_path)
    except Exception as exc:
        logger.warning("Could not execute SQL: %s", exc)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_daily(target_date: datetime, seed: int | None = None, dry_run: bool = False) -> None:
    """Generate and push all daily data for *target_date*."""
    day_seed = seed or int(target_date.strftime("%Y%m%d"))
    rng = random.Random(day_seed)

    customer_ids, product_ids = _load_reference_data()

    # Mid-sized e-commerce: ~50-80 orders/day, ~200-500 sessions, ~15-30 tickets
    num_orders = rng.randint(50, 80)
    num_sessions = rng.randint(200, 500)
    num_tickets = rng.randint(15, 30)

    logger.info("=== Daily data for %s (seed=%d) ===", target_date.date(), day_seed)
    logger.info("  Orders: %d | Sessions: %d | Tickets: %d", num_orders, num_sessions, num_tickets)

    # 1. Orders
    orders, items = generate_daily_orders(rng, target_date, customer_ids, product_ids, num_orders)
    sql_path = write_daily_sql(orders, items, target_date)
    if not dry_run:
        execute_daily_sql(sql_path)

    # 2. Clickstream events
    events = generate_daily_events(rng, target_date, customer_ids, product_ids, num_sessions)
    write_daily_events(events, target_date)
    if not dry_run:
        push_events_to_api(events)

    # 3. CS tickets
    tickets = generate_daily_tickets(rng, target_date, customer_ids, num_tickets)
    write_daily_tickets(tickets, target_date)

    logger.info("=== Daily generation complete for %s ===", target_date.date())


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate daily incremental data for Airflow ETL")
    parser.add_argument("--date", type=str, default=None,
                        help="Target date (YYYY-MM-DD). Default: today")
    parser.add_argument("--seed", type=int, default=None, help="Random seed (default: date as int)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Generate files only, don't push to services")
    args = parser.parse_args()

    if args.date:
        target_date = datetime.strptime(args.date, "%Y-%m-%d")
    else:
        target_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    run_daily(target_date, seed=args.seed, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
