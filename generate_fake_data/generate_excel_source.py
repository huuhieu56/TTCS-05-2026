"""Tạo dữ liệu nguồn Excel — CS_Tickets.xlsx từ Olist reviews."""

from __future__ import annotations

import argparse
import logging
import random
import sys
from datetime import datetime, timedelta
from pathlib import Path

from generate_fake_data.helpers import (
    deterministic_email,
    read_csv as _read_csv,
    shift_timestamp,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("generate_excel")

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_OUTPUT_DIR = _PROJECT_ROOT / "data_source" / "excel"

ISSUE_TYPES = ["Product Issue", "General Inquiry", "Positive Feedback"]
STATUSES = ["Open", "Resolved"]


def _derive_issue_type(rating: int) -> str:
    """rating ≤2 → Product Issue, =3 → General Inquiry, ≥4 → Positive Feedback."""
    if rating <= 2:
        return "Product Issue"
    elif rating == 3:
        return "General Inquiry"
    else:
        return "Positive Feedback"


def generate_cs_tickets(rng: random.Random, sample_frac: float = 1.0) -> None:
    """Tạo CS_Tickets.xlsx từ olist_order_reviews."""
    logger.info("=== Tạo CS_Tickets.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        logger.error("Cần cài openpyxl: pip install openpyxl")
        sys.exit(1)

    reviews = _read_csv("olist_order_reviews_dataset.csv")
    logger.info("  Đọc được %d reviews", len(reviews))

    # Map order → customer → customer_unique_id
    orders = _read_csv("olist_orders_dataset.csv")
    order_to_customer: dict[str, str] = {}
    for row in orders:
        order_to_customer[row["order_id"]] = row.get("customer_id", "")

    customers = _read_csv("olist_customers_dataset.csv")
    cid_to_uid: dict[str, str] = {}
    for row in customers:
        cid_to_uid[row["customer_id"]] = row["customer_unique_id"]

    if sample_frac < 1.0:
        k = max(1, int(len(reviews) * sample_frac))
        reviews = rng.sample(reviews, k)
        logger.info("  Lấy mẫu: %d reviews", len(reviews))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CS_Tickets"

    headers = [
        "Ticket_ID", "Order_ID", "Customer_Email",
        "Issue_Type", "Status", "Customer_Rating", "Reported_Date",
    ]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ticket_count = 0
    for i, row in enumerate(reviews, 1):
        review_id = row.get("review_id", f"rev_{i}")
        order_id = row.get("order_id", "")
        customer_id = order_to_customer.get(order_id, "")
        uid = cid_to_uid.get(customer_id, customer_id)

        try:
            rating = int(row.get("review_score", "3"))
            rating = max(1, min(5, rating))
        except (ValueError, TypeError):
            rating = 3

        ticket_id = review_id
        email = deterministic_email(uid) if uid else ""

        # ~5% email cố ý sai format (dirty data cho Spark xử lý)
        if email and rng.random() < 0.05:
            err = rng.choice(["no_at", "typo_domain", "no_dot"])
            if err == "no_at":
                email = email.replace("@", "")
            elif err == "typo_domain":
                email = (email.replace("gmail.com", "gmal.com")
                              .replace("hotmail.com", "hotmal.com")
                              .replace("outlook.com", "outlok.com")
                              .replace("yahoo.com.br", "yaho.com.br"))
            else:
                email = email.replace(".com", "com")

        issue_type = _derive_issue_type(rating)

        # Có answer_timestamp → Resolved, không → Open
        answer_ts = row.get("review_answer_timestamp", "").strip()
        status = "Resolved" if answer_ts else "Open"

        reported = row.get("review_creation_date",
                           row.get("review_answer_timestamp", "2018-01-01 00:00:00"))
        reported = shift_timestamp(reported)

        ws.append([ticket_id, order_id, email, issue_type, status, rating, reported])
        ticket_count += 1

    # Auto-fit column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = _OUTPUT_DIR / "CS_Tickets.xlsx"
    wb.save(output_path)
    logger.info("  Đã ghi %d tickets → %s", ticket_count, output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Tạo CS_Tickets.xlsx từ Olist reviews")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    parser.add_argument("--sample-frac", type=float, default=1.0, help="Tỉ lệ lấy mẫu (0.0–1.0)")
    args = parser.parse_args()

    rng = random.Random(args.seed)
    logger.info("Bắt đầu tạo dữ liệu Excel (seed=%d, sample=%.0f%%)",
                args.seed, args.sample_frac * 100)

    generate_cs_tickets(rng, args.sample_frac)
    logger.info("Hoàn tất! File đầu ra tại: %s", _OUTPUT_DIR)


if __name__ == "__main__":
    main()
